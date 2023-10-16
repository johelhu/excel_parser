#!/usr/bin/env python 

import openpyxl
import pandas as pd # importar libreria pandas, se usa para interactuar con tablas de datos

def inicializar_datos(archivo): # inicia la funcion 
    try: # intenta hacer
        # lee el excel llamado archivo, y convierte el ID en un str, todo lo guarda en dataframe(df)
        df = pd.read_excel(archivo, converters={'id_paciente': str}) 
        print('Se encontró el archivo de datos')

    except FileNotFoundError: # excepto si 
        # si no se encuentra el data entonces imprima el error y salga del programa
        print('No se encontró el archivo de datos')
        exit(1)

    return df # se retorna el valor y acaba la función en este caso nos guarda que que guardamos en df para usarlo en otro lugar

def inicializar_hojas(archivo):
    # Cargar el archivo Excel
    workbook = openpyxl.load_workbook('pacientes.xlsx')
    # Obtener los nombres de las hojas
    print("Nombres de las hojas en el archivo:")
    for sheet_name in workbook.sheetnames:
        print(sheet_name)
        
def mover_datos(archivo, id_paciente, origen, destino):
    # Cargar el archivo Excel
    workbook = openpyxl.load_workbook(archivo)

    # Determinar las hojas de origen y destino
    sheet_origen = workbook[origen]
    sheet_destino = workbook[destino]

    # Imprimir el ID que estamos buscando
    print("ID a buscar:", id_paciente)

    # Imprimir los IDs en la hoja de origen para verificar
    print("IDs en la hoja de origen:")
    for fila in sheet_origen.iter_rows(min_row=2, max_row=sheet_origen.max_row, max_col=1, values_only=True):
        print(fila[0])

    # Encontrar la fila con el ID especificado
    fila = 2  # Empezamos desde la segunda fila
    while str(sheet_origen.cell(row=fila, column=3).value) != id_paciente:
        fila += 1

    # Obtener los datos del paciente
    datos_paciente = [cell.value for cell in sheet_origen[fila]]

    # Encontrar la última fila vacía en la hoja de destino
    ultima_fila = sheet_destino.max_row + 1

    # Escribir los datos en la última fila vacía de la hoja de destino
    for i, valor in enumerate(datos_paciente):
        sheet_destino.cell(row=ultima_fila, column=i+1).value = valor

    # Eliminar los datos de la hoja de origen
    for cell in sheet_origen[fila]:
        cell.value = None

    # Guardar el archivo Excel
    workbook.save(archivo)
